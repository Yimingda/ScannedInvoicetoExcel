"""按模板配置把解析结果写入 Excel。

- 若模板文件存在：打开它，从 start_row 起按 columns 映射逐行写入，保留模板原有格式/表头。
- 若模板不存在：自动生成一个带中文表头的默认模板到 template.path，再写入。
把你的真实模板放到该路径、并在 config.yaml 的 columns 里改成对应列即可。
"""
from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string

# 字段 -> 默认表头中文名（仅在自动生成模板时使用）
_HEADER_LABELS = {
    "source_file": "来源文件",
    "invoice_date": "发票日期",
    "total_incl_tax": "带税总金额",
    "currency": "币种",
    "subtotal": "税前小计",
    "tax": "税额",
    "tip": "小费/服务费",
    "invoice_type": "类型",
    "doc_kind": "单据种类",
    "confidence": "置信度",
    "notes": "备注",
}


def _ensure_template(path: Path, sheet: str, header_row: int,
                     columns: Dict[str, str]) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet or "Sheet1"
    for field, col in columns.items():
        label = _HEADER_LABELS.get(field, field)
        ws[f"{col}{header_row}"] = label
    wb.save(str(path))


def build_workbook(records: List[Dict[str, Any]], template_cfg: Dict[str, Any],
                   template_source=None):
    """按模板把记录填好，返回 openpyxl Workbook（不落盘）。

    template_source: 可传入一个文件路径或 file-like（如上传的模板字节流）；
    留空则用 template_cfg['path']。
    """
    sheet = template_cfg.get("sheet") or None
    header_row = int(template_cfg.get("header_row", 1))
    start_row = int(template_cfg.get("start_row", header_row + 1))
    columns: Dict[str, str] = template_cfg["columns"]

    # 列映射的值支持 单列("B") 或 多列(["B","C"])——如开始/结束日期填同一天
    norm_columns: Dict[str, list] = {
        f: (v if isinstance(v, list) else [v]) for f, v in columns.items()
    }

    if template_source is None:
        path = Path(template_cfg["path"])
        _ensure_template(path, sheet or "Sheet1", header_row,
                         {f: cols[0] for f, cols in norm_columns.items()})
        template_source = str(path)

    wb = load_workbook(template_source)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    row = start_row
    for rec in records:
        for field, cols in norm_columns.items():
            val = rec.get(field)
            if val is None:
                continue
            if field == "doc_kind":
                val = {"invoice": "发票", "card_slip": "刷卡小票"}.get(val, val)
            for col in cols:
                ci = column_index_from_string(col)
                ws.cell(row=row, column=ci, value=val)
        row += 1
    return wb


def write_records(records: List[Dict[str, Any]], template_cfg: Dict[str, Any],
                  output_path: Path) -> None:
    wb = build_workbook(records, template_cfg)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
