"""跑完整流程并产出两个文件：
  1) 导入文件：只填模板的 B/C/K 列，可直接上传报销系统
  2) 复核文件：列出每条记录的全部字段+置信度+备注，供人工核对（尤其 low 置信度行）
用法： .venv\\Scripts\\python.exe finalize.py
"""
from __future__ import annotations
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from pathlib import Path
import warnings; warnings.filterwarnings("ignore")
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage

from invoicescanner import pipeline, excel_writer

ROOT = Path(__file__).parent
cfg = yaml.safe_load(open(ROOT / "config.yaml", encoding="utf-8"))

out_dir = ROOT / cfg.get("output_dir", "output")
out_dir.mkdir(exist_ok=True)
cfg["_crop_dir"] = str(out_dir / "crops")   # 让 pipeline 存下每张小票的裁切图

print("开始处理 input/ 下的文件……")
records = pipeline.process_all(cfg)
if not records:
    print("无记录"); sys.exit(1)

# 1) 导入文件
import_path = out_dir / "报销导入_已填.xlsx"
excel_writer.write_records(records, cfg["template"], import_path)
print(f"[导入文件] {import_path}  （{len(records)} 行，B/C=日期 K=金额）")

# 2) 复核文件（内嵌每张小票裁切图，对着原图补录）
review = Workbook(); ws = review.active; ws.title = "复核"
cols = ["来源(页-区)", "日期", "金额", "币种", "税前", "税", "小费",
        "种类", "类型", "置信度", "备注", "原图（对着改）"]
ws.append(cols)
for c in ws[1]:
    c.font = Font(bold=True)
red = PatternFill("solid", fgColor="FFCDD2")
yellow = PatternFill("solid", fgColor="FFF9C4")
IMG_COL = 12   # L 列放图
for r in records:
    kind = {"invoice": "发票", "card_slip": "刷卡小票"}.get(r.get("doc_kind"), "")
    ws.append([r.get("source_file"), r.get("invoice_date"), r.get("total_incl_tax"),
               r.get("currency"), r.get("subtotal"), r.get("tax"), r.get("tip"),
               kind, r.get("invoice_type"), r.get("confidence"), r.get("notes")])
    rownum = ws.max_row
    row = ws[rownum]
    conf = r.get("confidence")
    if conf == "low" or not r.get("invoice_date") or r.get("total_incl_tax") is None:
        for c in row: c.fill = red
    elif conf == "medium":
        for c in row: c.fill = yellow
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
    # 内嵌裁切图，并把行高设成图高，便于对照
    cp = r.get("_crop_path")
    if cp and Path(cp).exists():
        try:
            img = XLImage(cp)
            scale = min(1.0, 300 / img.width)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            ws.row_dimensions[rownum].height = max(60, img.height * 0.75)
            ws.add_image(img, f"L{rownum}")
        except Exception:
            pass
widths = [20, 12, 10, 6, 8, 7, 7, 9, 9, 8, 46, 46]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+i)].width = w
ws.freeze_panes = "A2"
review_path = out_dir / "复核表.xlsx"
review.save(review_path)
low = sum(1 for r in records if r.get("confidence") == "low"
          or not r.get("invoice_date") or r.get("total_incl_tax") is None)
print(f"[复核文件] {review_path}  （红色={low} 行需重点核对，L列可对着原图改）")
